import openpyxl
from openpyxl import Workbook
import os


def write(args):
    result = dict()
    with open(args.filename, 'r', encoding='utf-8') as fp:
        lines = fp.readlines()

    for line in lines:
        tokens = line.strip().split()
        language = tokens[0].split(':')[-1]

        result[language] = result.get(language, dict())

        feature_index = int(tokens[1].split(':')[-1])
        value = float(tokens[3].split(':')[-1])
        maximum = result[language].get(feature_index, -0.01)
        if value > maximum:
            result[language][feature_index] = value

    # excel
    path = './layer ' + str(args.layer)
    if not os.path.exists(path):
        os.makedirs(path)
    fn = path + '/layer' + str(args.layer) + '_result_' + args.model + '.xlsx'
    wb = Workbook()
    wb.save(fn)
    wb = openpyxl.load_workbook(fn)
    ws = wb.worksheets[0]

    for i in range(len(args.features)):
        _ = ws.cell(row=i + 2, column=1, value=args.features[i])

    for i in range(len(args.languages)):
        _ = ws.cell(row=1, column=2 + i, value=args.languages[i])

    for j in range(len(args.languages)):
        language = args.languages[j]
        indexs = list(result[language].keys())
        for i in range(len(args.features)):
            if i in indexs:
                _ = ws.cell(row=i + 2, column=j + 2, value=result[language][i])
            else:
                _ = ws.cell(row=i + 2, column=j + 2, value='/')
    wb.save(fn)
